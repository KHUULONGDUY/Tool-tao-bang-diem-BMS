import pandas as pd
import numpy as np
import shutil
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font
import math 

global a1,a2,a3,a4,a5,a6,a7
A1=[12,26,29,30,31,34,35,46,49]
A2=[13,15,16,17,18,20,21,22,27,37,38,39,41,42,43,44,48]
A3=[14,19,23,24,25,28,32,33,36,40,45,47]
rowNum = 7

save_path, file_out= "",""

cols= {'Name':0,'AHU':1,'AHU_name':2,'BMS':4,'AHU_name2':6,'AI_EMS':7,'AI_BMS':8,'AO_BMS':9,'DI_BMS':10,'DO_BMS':11,'Fan_AI':12,'Valve_AI':13,'DPT_filter':18,
       'Room_filter':19,'Room_RH':24,'Room_pres':25,'Fan_AO':26,'Valve_AO':27,'Fan_run':29,'Fan_trip':30,'Comp_run':32,'Comp_trip':33,
       'Res_power':34,'Res_error':35,'Fan_DO':46,'Comp_DO':47,'Res_DO':49, 'wind_AI':14, 'wind_AO':28, 'air_switch':38, 'temp_switch':39,
       'Wire_A1':53,'Wire_A2':54,'Wire_A3':55,'A1_coef':56, 'A7_coef':57, 'pvc20':58, 'AHU_temp':15, 'AHU_RH':17,'DPS_filter':41}

def CopyTemplateFile(file_path,save_path,DA_name):
    global file_out, rowNum 
    rowNum = 7
    #file_path= r"C:\Users\84935\source\repos\PointDetail\PointDetail\Template.xlsx"
    file_out= save_path + "\\" + DA_name + "_BMS_KL_detail.xlsx"
    #file_out = savepath + Name
    shutil.copyfile(file_path, file_out)
    return None
    
def GetCols(file_path):
    #file_path= r"C:\Users\84935\source\repos\PointDetail\PointDetail\Template.xlsx"
    xlsx= pd.ExcelFile(file_path)
    sheet_list= xlsx.sheet_names
    df= pd.read_excel(file_path, sheet_name= sheet_list[0])
    xlsx.close()
    col_list= df.columns
    del df
    return col_list

def change_value(A1,A2,A3,A4,A5,A6,A7):
    global a1,a2,a3,a4,a5,a6,a7
    a1 = A1
    a2 = A2
    a3 = A3
    a4 = A4
    a5 = A5
    a6 = A6
    a7 = A7
def Wiring_sum(col_list, row_df):
    AI_BMS=0
    AI_EMS=0
    AO_BMS=0
    DI_BMS=0
    DO_BMS=0
    sum_A1= 0
    sum_A2=0
    sum_A3=0
    for i in range(12, len(col_list)-9):
        # Sum AI of BMS
        if i in range(12,24):
            if math.isnan(row_df[i]):
                AI_BMS +=0
            else:
                AI_BMS += row_df[i]
        # Sum AI of EMS
        if i==24 or i==25:
            if math.isnan(row_df[i]):
                AI_EMS +=0
            else:
                AI_EMS += row_df[i]
        # Sum AO of BMS
        if i in range(26,29):
            if math.isnan(row_df[i]):
                AO_BMS +=0
            else:
                AO_BMS += row_df[i]
        # Sum DI of BMS
        if i in range(29,46):
            if math.isnan(row_df[i]):
                DI_BMS +=0
            else:
                DI_BMS += row_df[i]
        # Sum DO of BMS
        if i in range(46,50):
            if math.isnan(row_df[i]):
                DO_BMS +=0
            else:
                DO_BMS += row_df[i]
        # Sum of A1
        if i in A1:
            if math.isnan(row_df[i]):
                sum_A1 += 0
            else:
                sum_A1 += row_df[i]  
        # Sum of A2
        elif i in A2:
            if math.isnan(row_df[i]):
                sum_A2 += 0
            else:
                sum_A2 += row_df[i]
        # Sum of A3
        elif i in A3:
            if math.isnan(row_df[i]):
                sum_A3 += 0
            else:
                sum_A3 += row_df[i]
        else:
            pass
    # Put data in the correct position
    row_df[cols['AI_EMS']]= AI_EMS
    row_df[cols['AI_BMS']]= AI_BMS
    row_df[cols['AO_BMS']]= AO_BMS
    row_df[cols['DI_BMS']]= DI_BMS
    row_df[cols['DO_BMS']]= DO_BMS
    row_df[cols['Wire_A1']]= sum_A1*a1
    row_df[cols['Wire_A2']]= sum_A2*a2
    row_df[cols['Wire_A3']]= sum_A3*a3
    row_df[cols['A1_coef']]= a1
    row_df[cols['A7_coef']]= a7
    row_df[cols['pvc20']]= row_df[cols['Wire_A1']]*a4 + row_df[cols['Wire_A2']]*a5 + row_df[cols['Wire_A3']]*a6
    return row_df


def AddData(col_list,AHU_name, VSD_fan, NoVSD_fan, hot_coil, cool_coil, compressor, resistor, DPT_filter,DPS_filter, Room_filter, Room_RH,Room_pres,wind_valve):
    row_df= len(col_list)*[np.nan]
    row_df[cols['Name']]= 'Air Handing Unit'
    row_df[cols['AHU']]= 'AHU'
    row_df[cols['AHU_name']]= AHU_name
    row_df[cols['BMS']]= 'BMS'
    row_df[cols['AHU_name2']]= "DDC-"+AHU_name
    # For fan 
    row_df[cols['Fan_AI']]= VSD_fan
    row_df[cols['Fan_AO']]= VSD_fan
    row_df[cols['Fan_run']]= VSD_fan + NoVSD_fan 
    row_df[cols['Fan_trip']]= VSD_fan + NoVSD_fan 
    row_df[cols['Fan_DO']]= VSD_fan + NoVSD_fan 
    # For valve of hot coil and cool coil 
    valve=0
    if hot_coil== True:
        valve +=1
    if cool_coil== True:
        valve +=1
    row_df[cols['Valve_AI']]= valve
    row_df[cols['Valve_AO']]= valve
    # For AHU sensor
    if (cool_coil== True or compressor >1) and (hot_coil== False and resistor==0):
        row_df[cols['AHU_temp']]= 1
    if (cool_coil== True or compressor >1) and (hot_coil== True or resistor==1):
        row_df[cols['AHU_RH']]= 1
    # For room sensors
    row_df[cols['Room_filter']]= Room_filter
    row_df[cols['Room_RH']]= Room_RH
    row_df[cols['Room_pres']]= Room_pres
    # For resistor
    row_df[cols['Res_power']]= resistor
    row_df[cols['Res_error']]= resistor
    row_df[cols['Res_DO']]= resistor
    if resistor >=1:
        row_df[cols['air_switch']]=1
        row_df[cols['temp_switch']]=1
    # For compressor
    row_df[cols['Comp_run']]= compressor
    row_df[cols['Comp_trip']]= compressor
    row_df[cols['Comp_DO']]= compressor
    # For AHU filter
    row_df[cols['DPT_filter']]= DPT_filter
    row_df[cols['DPS_filter']]= DPS_filter
    # For wind valve
    row_df[cols['wind_AI']]= wind_valve
    row_df[cols['wind_AO']]= wind_valve
    row_df= Wiring_sum(col_list, row_df)
    return row_df


def Set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    return None

def Align(ws, rowNum, length):
    ex_list=[0,1,2,3,4,7]
    for i in range(1,length+1):
        ws.cell(rowNum, i).font= Font(size=10)
        if i in ex_list:
            continue
        ws.cell(rowNum, i).alignment= Alignment(horizontal="center", vertical='center')
    return None

def WireSheet():
    work_book= openpyxl.load_workbook(file_out)
    ws2= work_book.worksheets[1]
    ws2.cell(2,2).value= a1
    ws2.cell(3,2).value= a2
    ws2.cell(4,2).value= a3
    ws2.cell(5,2).value= a4
    ws2.cell(6,2).value= a5
    ws2.cell(7,2).value= a6
    ws2.cell(8,2).value= a7
    work_book.save(file_out) 
    return None
def change_value(A1,A2,A3,A4,A5,A6,A7):
    global a1,a2,a3,a4,a5,a6,a7
    a1 = A1
    a2 = A2
    a3 = A3
    a4 = A4
    a5 = A5
    a6 = A6
    a7 = A7

def DataHandling(col_list,AHU_name, VSD_fan, NoVSD_fan, hot_coil, cool_coil, compressor, resistor, DPT_filter,DPS_filter, Room_filter, Room_RH,Room_pres,wind_valve):
    global rowNum
    row_df= AddData(col_list,AHU_name, VSD_fan, NoVSD_fan, hot_coil, cool_coil, compressor, resistor, DPT_filter,DPS_filter, Room_filter, Room_RH,Room_pres,wind_valve)
    work_book= openpyxl.load_workbook(file_out)
    ws= work_book.worksheets[0]
    ws.append(row_df)
    rowNum +=1
    cell_range= 'A'+str(rowNum)+':BG'+str(rowNum)
    Set_border(ws,cell_range)
    Align(ws,rowNum, len(col_list))
    #work_book.save(r"C:\Users\84935\source\repos\PointDetail\PointDetail\Template1.xlsx") 
    work_book.save(file_out) 
    del row_df
    return None

def get_col_list(file_path,save_path,DA_name):
    CopyTemplateFile(file_path,save_path,DA_name)
    col_list= GetCols(file_path)
    return col_list
   